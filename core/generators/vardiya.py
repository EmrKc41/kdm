"""FONKSİYON 3 — Vardiya listesi üreticisi (sıfırdan tasarım).

Bu fonksiyonun şablonu yoktur ve çizim nesnesi içermez; bu yüzden burada
openpyxl ile YAZMAK güvenli ve doğru araçtır. Fonksiyon 1 ve 2'deki
ZIP/XML cerrahisi yalnızca şablon sadakati gerektiği için kullanılıyordu.

YERLEŞİM
    A, B ve C vardiyaları TEK sayfada YAN YANA üç blok halinde dizilir.
    Sayfa adı tarihten türetilir: 17 Ağustos -> "Ağustos3.Hafta".

    Yan yana yerleşim seçildiği için otomatik filtre uygulanmaz: Excel bir
    sayfada yalnızca tek bir filtre aralığı destekler ve üç bloktan birine
    bağlanması yanıltıcı olurdu. Dondurulmuş başlıklar korunur.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..errors import GirdiHatasi
from ..models import (
    HaftalikVardiya,
    VardiyaKaydi,
    VardiyaVerisi,
    vardiya_harfi,
)
from ..rules import KuralMotoru, varsayilan_kurallar
from ..tarih import hafta_adi

# --- Kurumsal görünüm --------------------------------------------------------

RENK_BANT = "1F3864"           # sütun başlığı bandı — koyu lacivert
RENK_BANT_YAZI = "FFFFFF"
RENK_BASLIK_BLOGU = "2E5395"
RENK_ALTERNATIF = "EEF2F8"
RENK_KENARLIK = "BFC9DA"

#: Her vardiyanın kendi bant rengi — bloklar bir bakışta ayrışsın.
VARDIYA_RENKLERI = {
    "A": "1F3864",
    "B": "1E6B4F",
    "C": "7A3E12",
}

FONT_ADI = "Calibri"

SUTUNLAR = [
    ("sira", "SIRA", 6),
    ("ad_soyad", "AD SOYAD", 24),
    ("unvan", "ÜNVAN", 20),
    ("calisma_yeri", "ÇALIŞACAĞI YER / HAT", 24),
    ("telefon", "TELEFON NO", 16),
    ("durak", "DURAK İSMİ", 20),
]

BLOK_SUTUN_SAYISI = len(SUTUNLAR)
BLOK_ARALIGI = 1               # bloklar arasındaki dar boşluk sütunu
ARALIK_GENISLIGI = 2.5

GENEL_BASLIK_SATIRI = 1
TARIH_SATIRI = 2
VARDIYA_BANT_SATIRI = 4
BASLIK_SATIRI = 5              # sütun başlıklarının satırı
ILK_VERI_SATIRI = BASLIK_SATIRI + 1

#: Liste kısa olsa bile elle doldurmak için bırakılan yedek satır sayısı.
YEDEK_SATIR = 3

MAKS_KAYIT = 5000

_ince = Side(style="thin", color=RENK_KENARLIK)
KENARLIK = Border(left=_ince, right=_ince, top=_ince, bottom=_ince)


def blok_ilk_sutunu(index: int) -> int:
    """0 tabanlı blok indeksinin başladığı sütun (1 tabanlı)."""
    return 1 + index * (BLOK_SUTUN_SAYISI + BLOK_ARALIGI)


def uret(
    veri: HaftalikVardiya | VardiyaVerisi,
    motor: KuralMotoru | None = None,
) -> Workbook:
    """Üç vardiyayı tek sayfada yan yana içeren çalışma kitabı üretir.

    Geriye dönük uyum: tek bir `VardiyaVerisi` verilirse yalnızca o vardiya
    kendi bloğunda yer alır, diğer iki blok boş kalır.
    """
    if isinstance(veri, VardiyaVerisi):
        # Vardiya harfi belirtilmemişse ilk bloğa yerleşir; kayıtlar
        # hiçbir koşulda sessizce düşürülmez.
        if not veri.vardiya_adi.strip():
            veri.vardiya_adi = "A"
        veri = HaftalikVardiya(
            tarih=veri.tarih,
            vardiyalar=[veri],
            normal_unvanlar=veri.normal_unvanlar,
        )
    _dogrula(veri)

    if motor is None:
        motor = KuralMotoru(varsayilan_kurallar(veri.normal_unvanlar))

    wb = Workbook()
    ws = wb.active
    ws.title = hafta_adi(veri.tarih)

    _yaz_genel_baslik(ws, veri)
    for i, vardiya in enumerate(veri.vardiyalar):
        _yaz_blok(ws, i, vardiya, motor)
    _bicimlendir_sayfa(ws, veri)
    return wb


def bos_sablon() -> Workbook:
    """Elle doldurulmak üzere boş haftalık çizelge.

    Her bloğun ilk satırında biçimi gösteren bir ÖRNEK kayıt bulunur.
    """
    ornek = VardiyaKaydi(
        ad_soyad="Ahmet Yılmaz",
        unvan="Kalite Operatörü",
        calisma_yeri="Pres Hattı 3",
        telefon="0532 000 00 00",
        durak="Merkez Durak",
    )
    veri = HaftalikVardiya(
        tarih="",
        vardiyalar=[
            VardiyaVerisi(vardiya_adi=h, kayitlar=[ornek]) for h in ("A", "B", "C")
        ],
    )
    wb = uret(veri)
    ws = wb.active

    satir = ILK_VERI_SATIRI + 1
    ws.cell(satir, blok_ilk_sutunu(0) + 1).value = (
        "↑ Örnek satırdır, silip kendi kayıtlarınızı yazın. Telefonlar metin "
        "biçimindedir, baştaki sıfır korunur. Ünvanı \"Kalite Operatörü\" "
        "olmayan satırlar otomatik kırmızı ve kalın gösterilir."
    )
    ws.cell(satir, blok_ilk_sutunu(0) + 1).font = Font(
        name=FONT_ADI, size=9, italic=True, color="606060"
    )
    return wb


# --- Bölümler ----------------------------------------------------------------


def _son_sutun(ws: Worksheet) -> int:
    return blok_ilk_sutunu(2) + BLOK_SUTUN_SAYISI - 1


def _yaz_genel_baslik(ws: Worksheet, veri: HaftalikVardiya) -> None:
    """Satır 1-2: tüm blokları kapsayan genel başlık ve tarih."""
    son = _son_sutun(ws)
    ws.merge_cells(start_row=GENEL_BASLIK_SATIRI, start_column=1,
                   end_row=GENEL_BASLIK_SATIRI, end_column=son)
    ws.merge_cells(start_row=TARIH_SATIRI, start_column=1,
                   end_row=TARIH_SATIRI, end_column=son)

    baslik = ws.cell(GENEL_BASLIK_SATIRI, 1)
    baslik.value = f"VARDİYA LİSTESİ — {hafta_adi(veri.tarih)}"
    baslik.font = Font(name=FONT_ADI, size=18, bold=True, color=RENK_BANT_YAZI)
    baslik.fill = PatternFill("solid", start_color=RENK_BASLIK_BLOGU)
    baslik.alignment = Alignment(horizontal="center", vertical="center")

    alt = ws.cell(TARIH_SATIRI, 1)
    alt.value = veri.tarih.strip()
    alt.font = Font(name=FONT_ADI, size=11, bold=True, color=RENK_BANT_YAZI)
    alt.fill = PatternFill("solid", start_color=RENK_BANT)
    alt.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[GENEL_BASLIK_SATIRI].height = 30
    ws.row_dimensions[TARIH_SATIRI].height = 20
    ws.row_dimensions[3].height = 8              # ince ayırıcı


def _yaz_blok(
    ws: Worksheet, index: int, vardiya: VardiyaVerisi, motor: KuralMotoru
) -> None:
    ilk = blok_ilk_sutunu(index)
    son = ilk + BLOK_SUTUN_SAYISI - 1
    renk = VARDIYA_RENKLERI.get(vardiya_harfi(vardiya.vardiya_adi) or "", RENK_BANT)

    # Vardiya bandı
    ws.merge_cells(start_row=VARDIYA_BANT_SATIRI, start_column=ilk,
                   end_row=VARDIYA_BANT_SATIRI, end_column=son)
    bant = ws.cell(VARDIYA_BANT_SATIRI, ilk)
    bant.value = " · ".join(
        p for p in (vardiya.baslik, vardiya.vardiya_saati.strip()) if p
    )
    bant.font = Font(name=FONT_ADI, size=13, bold=True, color=RENK_BANT_YAZI)
    bant.fill = PatternFill("solid", start_color=renk)
    bant.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[VARDIYA_BANT_SATIRI].height = 26

    # Sütun başlıkları
    for i, (_, etiket, _) in enumerate(SUTUNLAR):
        h = ws.cell(BASLIK_SATIRI, ilk + i)
        h.value = etiket
        h.font = Font(name=FONT_ADI, size=10, bold=True, color=RENK_BANT_YAZI)
        h.fill = PatternFill("solid", start_color=renk)
        h.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        h.border = KENARLIK
    ws.row_dimensions[BASLIK_SATIRI].height = 28

    _yaz_kayitlar(ws, ilk, vardiya, motor)


def _yaz_kayitlar(
    ws: Worksheet, ilk_sutun: int, vardiya: VardiyaVerisi, motor: KuralMotoru
) -> None:
    for sira, kayit in enumerate(vardiya.kayitlar, start=1):
        satir = ILK_VERI_SATIRI + sira - 1
        bicim = motor.bicim(kayit)

        font = Font(
            name=FONT_ADI,
            size=10,
            bold=bicim.kalin,
            italic=bicim.italik,
            color=bicim.renk,
        )
        golge = (
            PatternFill("solid", start_color=RENK_ALTERNATIF)
            if sira % 2 == 0
            else None
        )

        degerler = {
            "sira": sira,
            "ad_soyad": kayit.ad_soyad.strip(),
            "unvan": kayit.unvan.strip(),
            "calisma_yeri": kayit.calisma_yeri.strip(),
            "telefon": kayit.telefon.strip(),
            "durak": kayit.durak.strip(),
        }

        for i, (alan, _, _) in enumerate(SUTUNLAR):
            h = ws.cell(satir, ilk_sutun + i)
            h.value = degerler[alan]
            h.font = font
            h.border = KENARLIK
            if golge is not None:
                h.fill = golge

            if alan == "telefon":
                # Metin biçimi: baştaki sıfır kaybolmasın.
                h.number_format = "@"
                h.alignment = Alignment(horizontal="left", vertical="center")
            elif alan == "sira":
                h.alignment = Alignment(horizontal="center", vertical="center")
            else:
                h.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[satir].height = 17


def _bicimlendir_sayfa(ws: Worksheet, veri: HaftalikVardiya) -> None:
    son_satir = ILK_VERI_SATIRI + max(veri.en_uzun_liste, 1) + YEDEK_SATIR - 1
    son_sutun_harf = get_column_letter(_son_sutun(ws))

    # Blokların yüksekliği eşitlensin: kısa kalan bloklara boş çerçeve çizilir.
    _bos_cerceve_tamamla(ws, veri, son_satir)

    ws.freeze_panes = f"A{ILK_VERI_SATIRI}"
    # Otomatik filtre bilinçli olarak uygulanmıyor — bkz. modül başlığı.

    _sutun_genisliklerini_ayarla(ws, veri)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{son_sutun_harf}{son_satir}"
    ws.print_title_rows = f"{VARDIYA_BANT_SATIRI}:{BASLIK_SATIRI}"

    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.print_options.horizontalCentered = True

    ws.oddHeader.left.text = f"VARDİYA LİSTESİ - {hafta_adi(veri.tarih)}"
    ws.oddHeader.right.text = veri.tarih.strip()
    ws.oddFooter.right.text = "Sayfa &P / &N"
    ws.oddFooter.left.text = "&D &T"


def _bos_cerceve_tamamla(
    ws: Worksheet, veri: HaftalikVardiya, son_satir: int
) -> None:
    """Kısa kalan blokların altını boş ama çerçeveli satırlarla tamamlar."""
    for index, vardiya in enumerate(veri.vardiyalar):
        ilk = blok_ilk_sutunu(index)
        baslangic = ILK_VERI_SATIRI + len(vardiya.kayitlar)
        for satir in range(baslangic, son_satir + 1):
            sira = satir - ILK_VERI_SATIRI + 1
            golge = (
                PatternFill("solid", start_color=RENK_ALTERNATIF)
                if sira % 2 == 0
                else None
            )
            for i, (alan, _, _) in enumerate(SUTUNLAR):
                h = ws.cell(satir, ilk + i)
                h.border = KENARLIK
                h.font = Font(name=FONT_ADI, size=10)
                if golge is not None:
                    h.fill = golge
                if alan == "sira":
                    h.value = sira
                    h.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                elif alan == "telefon":
                    h.number_format = "@"
            ws.row_dimensions[satir].height = 17


def _sutun_genisliklerini_ayarla(ws: Worksheet, veri: HaftalikVardiya) -> None:
    """Aynı sütun üç blokta da AYNI genişlikte olur; bloklar hizalı görünür."""
    for i, (alan, etiket, taban) in enumerate(SUTUNLAR):
        en_uzun = len(etiket)
        for vardiya in veri.vardiyalar:
            for kayit in vardiya.kayitlar:
                deger = "" if alan == "sira" else str(getattr(kayit, alan, "") or "")
                en_uzun = max(en_uzun, len(deger))
        genislik = min(max(taban, en_uzun + 2), 40)

        for blok in range(3):
            harf = get_column_letter(blok_ilk_sutunu(blok) + i)
            ws.column_dimensions[harf].width = genislik

    for blok in range(1, 3):
        harf = get_column_letter(blok_ilk_sutunu(blok) - 1)
        ws.column_dimensions[harf].width = ARALIK_GENISLIGI


# --- Girdi doğrulaması -------------------------------------------------------


def _dogrula(veri: HaftalikVardiya) -> None:
    for vardiya in veri.vardiyalar:
        if len(vardiya.kayitlar) > MAKS_KAYIT:
            raise GirdiHatasi(
                f"{vardiya.baslik} en fazla {MAKS_KAYIT} kayıt alabilir "
                f"({len(vardiya.kayitlar)} kayıt girildi)."
            )
        for i, kayit in enumerate(vardiya.kayitlar, start=1):
            if not kayit.ad_soyad.strip():
                raise GirdiHatasi(
                    f"{vardiya.baslik} {i}. satırda Ad Soyad boş bırakılamaz."
                )

"""FONKSİYON 4 — Kalite uygunsuzluk takip raporu (sıfırdan tasarım).

Şablon dosyası yoktur; openpyxl ile yazılır (Fonksiyon 3 ile aynı yaklaşım).
A4 dikey, kurumsal başlık bandı ve düzeltici faaliyet tablosu içerir.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..errors import GirdiHatasi
from ..models import RaporSatiri, RaporVerisi

RENK_BASLIK = "1F3864"
RENK_BANT = "2E5395"
RENK_BANT_YAZI = "FFFFFF"
RENK_ALTERNATIF = "EEF2F8"
RENK_KENARLIK = "BFC9DA"
FONT_ADI = "Calibri"

SUTUNLAR = [
    ("sira", "SIRA", 5),
    ("tanim", "UYGUNSUZLUK TANIMI", 28),
    ("kok_neden", "KÖK NEDEN", 22),
    ("duzeltici_faaliyet", "DÜZELTİCİ FAALİYET", 24),
    ("sorumlu", "SORUMLU", 16),
    ("hedef_tarih", "HEDEF TARİH", 14),
    ("durum", "DURUM", 12),
]

BASLIK_SATIRI = 1
META_SATIRI = 2
OZET_BASLIK_SATIRI = 4
OZET_METIN_SATIRI = 5
TABLO_BASLIK_SATIRI = 7
ILK_VERI_SATIRI = 8

MAKS_SATIR = 100
YEDEK_SATIR = 5

_ince = Side(style="thin", color=RENK_KENARLIK)
KENARLIK = Border(left=_ince, right=_ince, top=_ince, bottom=_ince)


def uret(veri: RaporVerisi) -> Workbook:
    _dogrula(veri)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"

    _yaz_baslik(ws, veri)
    _yaz_meta(ws, veri)
    _yaz_ozet(ws, veri)
    _yaz_tablo_baslik(ws)
    _yaz_satirlar(ws, veri)
    _bicimlendir_sayfa(ws, veri)
    return wb


def bos_sablon() -> Workbook:
    ornek = RaporSatiri(
        tanim="Ölçü sapması — örnek satır, silip kendi kayıtlarınızı yazın",
        kok_neden="Ayar parametresi",
        duzeltici_faaliyet="Kalibrasyon ve yeniden ölçüm",
        sorumlu="Kalite Operatörü",
        hedef_tarih="GG.AA.YYYY",
        durum="Açık",
    )
    return uret(
        RaporVerisi(
            baslik="KALİTE UYGUNSUZLUK TAKİP RAPORU",
            konu="ORNEK-PARCA",
            rapor_no="R-001",
            ozet="Bu satır örnektir.",
            satirlar=[ornek],
        )
    )


def _son_sutun() -> int:
    return len(SUTUNLAR)


def _yaz_baslik(ws: Worksheet, veri: RaporVerisi) -> None:
    son = _son_sutun()
    ws.merge_cells(start_row=BASLIK_SATIRI, start_column=1, end_row=BASLIK_SATIRI, end_column=son)
    h = ws.cell(BASLIK_SATIRI, 1)
    h.value = veri.baslik.strip() or "KALİTE UYGUNSUZLUK TAKİP RAPORU"
    h.font = Font(name=FONT_ADI, size=16, bold=True, color=RENK_BANT_YAZI)
    h.fill = PatternFill("solid", start_color=RENK_BASLIK)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[BASLIK_SATIRI].height = 28


def _yaz_meta(ws: Worksheet, veri: RaporVerisi) -> None:
    etiketler = [
        ("Konu:", veri.konu),
        ("Rapor No:", veri.rapor_no),
        ("Tarih:", veri.tarih),
        ("Hazırlayan:", veri.hazirlayan),
        ("Genel Durum:", veri.genel_durum),
    ]
    ws.merge_cells(start_row=META_SATIRI, start_column=1, end_row=META_SATIRI, end_column=_son_sutun())
    metin = "   ".join(f"{e} {d}".strip() for e, d in etiketler if d or e)
    c = ws.cell(META_SATIRI, 1)
    c.value = metin.strip(" :")
    c.font = Font(name=FONT_ADI, size=10, bold=True, color=RENK_BANT_YAZI)
    c.fill = PatternFill("solid", start_color=RENK_BANT)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[META_SATIRI].height = 22


def _yaz_ozet(ws: Worksheet, veri: RaporVerisi) -> None:
    son = _son_sutun()
    ws.merge_cells(start_row=OZET_BASLIK_SATIRI, start_column=1, end_row=OZET_BASLIK_SATIRI, end_column=son)
    b = ws.cell(OZET_BASLIK_SATIRI, 1)
    b.value = "ÖZET"
    b.font = Font(name=FONT_ADI, size=10, bold=True)
    b.fill = PatternFill("solid", start_color=RENK_ALTERNATIF)

    ws.merge_cells(start_row=OZET_METIN_SATIRI, start_column=1, end_row=OZET_METIN_SATIRI, end_column=son)
    m = ws.cell(OZET_METIN_SATIRI, 1)
    m.value = veri.ozet.strip()
    m.font = Font(name=FONT_ADI, size=10)
    m.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[OZET_METIN_SATIRI].height = 36


def _yaz_tablo_baslik(ws: Worksheet) -> None:
    for i, (_, etiket, _) in enumerate(SUTUNLAR, start=1):
        h = ws.cell(TABLO_BASLIK_SATIRI, i)
        h.value = etiket
        h.font = Font(name=FONT_ADI, size=10, bold=True, color=RENK_BANT_YAZI)
        h.fill = PatternFill("solid", start_color=RENK_BASLIK)
        h.border = KENARLIK
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[TABLO_BASLIK_SATIRI].height = 32


def _yaz_satirlar(ws: Worksheet, veri: RaporVerisi) -> None:
    dolu = [s for s in veri.satirlar if s.dolu]
    satir_sayisi = max(len(dolu) + YEDEK_SATIR, YEDEK_SATIR)

    for idx in range(satir_sayisi):
        satir = ILK_VERI_SATIRI + idx
        kayit = dolu[idx] if idx < len(dolu) else RaporSatiri()
        golge = (
            PatternFill("solid", start_color=RENK_ALTERNATIF)
            if idx % 2 == 1
            else None
        )
        for col, (alan, _, _) in enumerate(SUTUNLAR, start=1):
            h = ws.cell(satir, col)
            h.border = KENARLIK
            h.font = Font(name=FONT_ADI, size=10)
            h.alignment = Alignment(vertical="top", wrap_text=True)
            if golge is not None:
                h.fill = golge
            if alan == "sira":
                h.value = idx + 1 if idx < len(dolu) else idx + 1
                h.alignment = Alignment(horizontal="center", vertical="top")
            else:
                h.value = getattr(kayit, alan, "") or ""
        ws.row_dimensions[satir].height = 40


def _bicimlendir_sayfa(ws: Worksheet, veri: RaporVerisi) -> None:
    for i, (_, _, genislik) in enumerate(SUTUNLAR, start=1):
        ws.column_dimensions[get_column_letter(i)].width = genislik

    dolu = max(len([s for s in veri.satirlar if s.dolu]), 0)
    son_satir = ILK_VERI_SATIRI + max(dolu + YEDEK_SATIR, YEDEK_SATIR) - 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(ILK_VERI_SATIRI, 1)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(_son_sutun())}{son_satir}"
    ws.print_title_rows = f"{TABLO_BASLIK_SATIRI}:{TABLO_BASLIK_SATIRI}"

    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True

    baslik = veri.baslik.strip() or "Kalite Raporu"
    ws.oddHeader.left.text = baslik
    ws.oddHeader.right.text = veri.tarih.strip()
    ws.oddFooter.right.text = "Sayfa &P / &N"


def _dogrula(veri: RaporVerisi) -> None:
    if not veri.konu.strip():
        raise GirdiHatasi("Konu / parça referansı zorunludur.")
    dolu = [s for s in veri.satirlar if s.dolu]
    if len(dolu) > MAKS_SATIR:
        raise GirdiHatasi(
            f"En fazla {MAKS_SATIR} uygunsuzluk satırı girilebilir ({len(dolu)} satır)."
        )

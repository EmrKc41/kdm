"""FONKSİYON 4 — Kalite uygunsuzluk takip raporu testleri."""

from __future__ import annotations

import io

import openpyxl
import pytest

from core.errors import GirdiHatasi
from core.generators import rapor
from core.models import RaporSatiri, RaporVerisi


@pytest.fixture
def ornek_rapor():
    return RaporVerisi(
        baslik="KALİTE UYGUNSUZLUK TAKİP RAPORU",
        konu="10598-AG",
        rapor_no="R-2026-014",
        tarih="17.08.2026",
        hazirlayan="Ahmet Yılmaz",
        genel_durum="Devam Ediyor",
        ozet="Pres hattında ölçü sapması tespit edildi.",
        satirlar=[
            RaporSatiri(
                tanim="Delik çapı tolerans dışı",
                kok_neden="Matkap aşınması",
                duzeltici_faaliyet="Matkap değişimi ve 100% kontrol",
                sorumlu="Kalite Operatörü",
                hedef_tarih="20.08.2026",
                durum="Açık",
            ),
            RaporSatiri(
                tanim="Yüzey çizik",
                kok_neden="Taşıma sırasında sürtünme",
                duzeltici_faaliyet="Koruyucu aparat takıldı",
                sorumlu="Üretim Sorumlusu",
                hedef_tarih="18.08.2026",
                durum="Kapalı",
            ),
        ],
    )


def _ac(wb) -> openpyxl.Workbook:
    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return openpyxl.load_workbook(tampon)


def test_uretim_acilir(ornek_rapor):
    assert _ac(rapor.uret(ornek_rapor)).active.title == "Rapor"


def test_baslik_yazilir(ornek_rapor):
    ws = _ac(rapor.uret(ornek_rapor)).active
    assert "KALİTE UYGUNSUZLUK" in str(ws.cell(1, 1).value)


def test_satir_icerigi(ornek_rapor):
    ws = _ac(rapor.uret(ornek_rapor)).active
    assert ws.cell(rapor.ILK_VERI_SATIRI, 2).value == "Delik çapı tolerans dışı"
    assert ws.cell(rapor.ILK_VERI_SATIRI + 1, 7).value == "Kapalı"


def test_bos_konu_reddedilir():
    with pytest.raises(GirdiHatasi, match="Konu"):
        rapor.uret(RaporVerisi(konu=""))


def test_bos_sablon_ornek_satir_icerir():
    ws = _ac(rapor.bos_sablon()).active
    assert "örnek" in str(ws.cell(rapor.ILK_VERI_SATIRI, 2).value).lower()


def test_turkce_karakter_korunur(ornek_rapor):
    ws = _ac(rapor.uret(ornek_rapor)).active
    assert "ölçü" in str(ws.cell(rapor.OZET_METIN_SATIRI, 1).value).lower()

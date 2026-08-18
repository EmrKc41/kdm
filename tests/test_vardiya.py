"""FONKSİYON 3 — Vardiya listesi üreticisi ve kural motoru testleri."""

from __future__ import annotations

import io

import openpyxl
import pytest

from core.errors import GirdiHatasi
from core.generators import vardiya
from core.importers import csvden_oku, exceldan_oku, oku
from core.models import VardiyaKaydi, VardiyaVerisi
from core.rules import BicimKurali, KayitBicimi, KuralMotoru, varsayilan_kurallar


@pytest.fixture
def ornek_vardiya():
    return VardiyaVerisi(
        vardiya_adi="A",
        tarih="17.08.2026",
        kayitlar=[
            VardiyaKaydi("Ahmet Yılmaz", "Kalite Operatörü", "Pres Hattı 3",
                         "0532 111 22 33", "Merkez Durak"),
            VardiyaKaydi("Ayşe Çelik", "Kalite Sorumlusu", "Pres Hattı 3",
                         "0533 222 33 44", "Şehitler Durağı"),
            VardiyaKaydi("Mustafa Şahin", "Kalite Operatörü", "Kaynak Hattı 1",
                         "0534 333 44 55", "Merkez Durak"),
            VardiyaKaydi("Zeynep Öztürk", "Vardiya Formeni", "Kaynak Hattı 1",
                         "0535 444 55 66", "Sanayi Durağı"),
            VardiyaKaydi("Emre Doğan", "Kalite Mühendisi", "Tüm Hatlar",
                         "0536 555 66 77", "Merkez Durak"),
        ],
    )


@pytest.fixture
def haftalik():
    from core.models import HaftalikVardiya

    return HaftalikVardiya(
        tarih="17.08.2026",
        vardiyalar=[
            VardiyaVerisi(vardiya_adi="A", kayitlar=[
                VardiyaKaydi("A-Kisi", "Kalite Operatörü", "Pres 3", "0532 1", "Merkez"),
                VardiyaKaydi("A-Kisi2", "Vardiya Formeni", "Pres 3", "0533 2", "Merkez"),
            ]),
            VardiyaVerisi(vardiya_adi="B", kayitlar=[
                VardiyaKaydi("B-Kisi", "Kalite Operatörü", "Kaynak 1", "0534 3", "Sanayi"),
            ]),
            VardiyaVerisi(vardiya_adi="C", kayitlar=[
                VardiyaKaydi("C-Kisi", "Proses Şefi", "Tüm Hatlar", "0535 4", "Merkez"),
            ]),
        ],
    )


def _kaydet_ve_ac(wb) -> openpyxl.Workbook:
    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return openpyxl.load_workbook(tampon)


# --- 3.2 Biçimlendirme kuralı ------------------------------------------------


def test_kalite_operatoru_siyah_normal(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    # 1. ve 3. kayıt Kalite Operatörü
    for satir in (vardiya.ILK_VERI_SATIRI, vardiya.ILK_VERI_SATIRI + 2):
        f = ws.cell(satir, 2).font
        assert f.bold is False, f"{satir}. satır kalın olmamalı."
        assert f.color.rgb.endswith("000000"), f"{satir}. satır siyah olmalı."


def test_diger_unvanlar_kirmizi_kalin(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    # 2., 4., 5. kayıt: Sorumlu, Formen, Mühendis
    for offset in (1, 3, 4):
        satir = vardiya.ILK_VERI_SATIRI + offset
        f = ws.cell(satir, 2).font
        assert f.bold is True, f"{satir}. satır kalın olmalı."
        assert f.color.rgb.endswith("C00000"), f"{satir}. satır kırmızı olmalı."


def test_kural_motoru_genisletilebilir(ornek_vardiya):
    """Ünvan listesi genişletilince kural davranışı değişmeli."""
    ornek_vardiya.normal_unvanlar = ["Kalite Operatörü", "Kalite Sorumlusu"]
    motor = KuralMotoru(varsayilan_kurallar(ornek_vardiya.normal_unvanlar))
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya, motor)).active

    # Artık Kalite Sorumlusu da normal
    assert ws.cell(vardiya.ILK_VERI_SATIRI + 1, 2).font.bold is False
    # Formen hâlâ vurgulu
    assert ws.cell(vardiya.ILK_VERI_SATIRI + 3, 2).font.bold is True


def test_kural_motoru_sabit_kodlanmamis():
    """Kural motoru ünvanları bilmez; yalnızca veriyi değerlendirir."""
    motor = KuralMotoru(
        [
            BicimKurali(
                ad="Mühendisler mavi",
                alan="unvan",
                operator="icerir",
                degerler=["mühendis"],
                bicim=KayitBicimi(kalin=True, renk="0070C0"),
            )
        ]
    )
    assert motor.bicim(VardiyaKaydi("A", "Kalite Mühendisi")).renk == "0070C0"
    assert motor.bicim(VardiyaKaydi("B", "Kalite Operatörü")).renk == "000000"


def test_kural_turkce_buyuk_kucuk_harfe_duyarsiz():
    motor = KuralMotoru(varsayilan_kurallar(["Kalite Operatörü"]))
    for yazim in ("Kalite Operatörü", "KALİTE OPERATÖRÜ", "kalite operatörü"):
        assert motor.bicim(VardiyaKaydi("A", yazim)).kalin is False, yazim


def test_kural_kaydet_yukle(tmp_path):
    yol = tmp_path / "kurallar.json"
    KuralMotoru(varsayilan_kurallar(["Kalite Operatörü", "Stajyer"])).kaydet(yol)

    motor = KuralMotoru.yukle(yol)
    assert motor.bicim(VardiyaKaydi("A", "Stajyer")).kalin is False
    assert motor.bicim(VardiyaKaydi("B", "Şef")).kalin is True


def test_gecersiz_operator_anlasilir_hata():
    motor = KuralMotoru(
        [BicimKurali("Bozuk", "unvan", "her_neyse", ["x"], KayitBicimi())]
    )
    with pytest.raises(GirdiHatasi, match="tanınmayan bir kural operatörü"):
        motor.bicim(VardiyaKaydi("A", "B"))


# --- 3.3 Tasarım -------------------------------------------------------------


def test_telefon_metin_bicimi_ve_bastaki_sifir(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    h = ws.cell(vardiya.ILK_VERI_SATIRI, 5)
    assert h.number_format == "@", "Telefon metin biçiminde değil."
    assert h.value == "0532 111 22 33", "Baştaki sıfır kaybolmuş."
    assert isinstance(h.value, str)


def test_dondurulmus_baslik(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    assert ws.freeze_panes == f"A{vardiya.ILK_VERI_SATIRI}"


def test_otomatik_filtre_uygulanmaz(ornek_vardiya):
    """Yan yana 3 blokta tek filtre araligi yaniltici olurdu; bilincli olarak yok."""
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    assert ws.auto_filter.ref is None


def test_a4_yatay_baski_ayarlari(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    assert ws.page_setup.orientation == "landscape"
    assert str(ws.page_setup.paperSize) == "9"        # A4
    assert ws.print_area is not None


def test_tekrarlayan_baslik_satiri_ve_alt_bilgi(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    # openpyxl aralığı mutlak biçime normalize eder: "4:4" -> "$4:$4"
    assert ws.print_title_rows.replace("$", "") == (
        f"{vardiya.VARDIYA_BANT_SATIRI}:{vardiya.BASLIK_SATIRI}"
    )
    assert "&P" in ws.oddFooter.right.text, "Sayfa numarası alt bilgide yok."


def test_ust_bilgide_hafta_bilgisi(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    assert "Ağustos3.Hafta" in (ws.oddHeader.left.text or "")
    assert "17.08.2026" in (ws.oddHeader.right.text or "")


def test_genel_baslik_ve_vardiya_bantlari(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    assert ws.cell(1, 1).value == "VARDİYA LİSTESİ — Ağustos3.Hafta"
    assert "17.08.2026" in ws.cell(2, 1).value

    bant = ws.cell(vardiya.VARDIYA_BANT_SATIRI, vardiya.blok_ilk_sutunu(0)).value
    assert "A VARDİYASI" in bant and "24.00 - 08.00" in bant


def test_alternatif_satir_golgelendirmesi(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    tek = ws.cell(vardiya.ILK_VERI_SATIRI, 2)          # 1. kayıt
    cift = ws.cell(vardiya.ILK_VERI_SATIRI + 1, 2)     # 2. kayıt
    assert cift.fill.start_color.rgb.endswith(vardiya.RENK_ALTERNATIF)
    assert tek.fill.fill_type is None


def test_ince_kenarliklar_ve_otomatik_genislik(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    h = ws.cell(vardiya.ILK_VERI_SATIRI, 2)
    assert h.border.left.style == "thin"
    assert ws.column_dimensions["B"].width >= len("Mustafa Şahin")


def test_turkce_karakterler(ornek_vardiya):
    ws = _kaydet_ve_ac(vardiya.uret(ornek_vardiya)).active
    degerler = [ws.cell(vardiya.ILK_VERI_SATIRI + i, 2).value for i in range(5)]
    assert "Ayşe Çelik" in degerler
    assert "Zeynep Öztürk" in degerler


# --- 3.4 Boş şablon ----------------------------------------------------------


def test_bos_sablon_ornek_satir_ve_aciklama_iceriyor():
    ws = _kaydet_ve_ac(vardiya.bos_sablon()).active
    for blok in range(3):
        sutun = vardiya.blok_ilk_sutunu(blok) + 1
        assert ws.cell(vardiya.ILK_VERI_SATIRI, sutun).value == "Ahmet Yılmaz"

    aciklama = ws.cell(
        vardiya.ILK_VERI_SATIRI + 1, vardiya.blok_ilk_sutunu(0) + 1
    ).value
    assert aciklama and "Örnek satırdır" in aciklama


# --- 3.1 İçe aktarma ---------------------------------------------------------


def test_csv_ice_aktarma_noktali_virgul():
    veri = (
        "Ad Soyad;Ünvan;Çalışacağı Yer;Telefon No;Durak İsmi\n"
        "Ahmet Yılmaz;Kalite Operatörü;Pres 3;0532 111 22 33;Merkez\n"
        "Ayşe Çelik;Kalite Sorumlusu;Pres 3;0533 222 33 44;Şehitler\n"
    ).encode("utf-8")

    kayitlar = csvden_oku(veri)
    assert len(kayitlar) == 2
    assert kayitlar[0].ad_soyad == "Ahmet Yılmaz"
    assert kayitlar[1].unvan == "Kalite Sorumlusu"
    assert kayitlar[0].telefon == "0532 111 22 33"


def test_csv_virgul_ayraci_ve_esnek_basliklar():
    veri = "isim,gorev,hat,gsm,servis\nAli Vural,Şef,Kaynak,05321112233,Merkez\n"
    kayitlar = csvden_oku(veri.encode("utf-8"))
    assert kayitlar[0].ad_soyad == "Ali Vural"
    assert kayitlar[0].unvan == "Şef"
    assert kayitlar[0].calisma_yeri == "Kaynak"
    assert kayitlar[0].durak == "Merkez"


def test_csv_turkce_kodlama_cp1254():
    veri = "Ad Soyad;Ünvan\nAyşe Çelik;Şef\n".encode("cp1254")
    kayitlar = csvden_oku(veri)
    assert kayitlar[0].ad_soyad == "Ayşe Çelik"
    assert kayitlar[0].unvan == "Şef"


def test_ad_soyad_sutunu_yoksa_anlasilir_hata():
    with pytest.raises(GirdiHatasi, match="Ad Soyad"):
        csvden_oku(b"Kolon1;Kolon2\na;b\n")


def test_excel_ice_aktarma(ornek_vardiya):
    """Üretilen bir vardiya listesi tekrar içe aktarılabilmeli."""
    tampon = io.BytesIO()
    vardiya.uret(ornek_vardiya).save(tampon)
    tampon.seek(0)

    # Başlık bloğu satırları atlanmalı; okuyucu başlık satırını bulmalı.
    wb = openpyxl.load_workbook(tampon)
    ws = wb.active
    for _ in range(vardiya.BASLIK_SATIRI - 1):
        ws.delete_rows(1)
    yeni = io.BytesIO()
    wb.save(yeni)

    kayitlar = exceldan_oku(yeni.getvalue())
    assert len(kayitlar) == 5
    assert kayitlar[0].ad_soyad == "Ahmet Yılmaz"


def test_desteklenmeyen_dosya_tipi():
    with pytest.raises(GirdiHatasi, match="desteklenmeyen bir dosya tipi"):
        oku("liste.pdf", b"%PDF")


# --- Sayfa adı: ayın kaçıncı haftası -----------------------------------------


@pytest.mark.parametrize(
    "tarih, beklenen",
    [
        ("17.08.2026", "Ağustos3.Hafta"),
        ("01.08.2026", "Ağustos1.Hafta"),
        ("07.08.2026", "Ağustos1.Hafta"),
        ("08.08.2026", "Ağustos2.Hafta"),
        ("2026-08-17", "Ağustos3.Hafta"),
        ("01.01.2027", "Ocak1.Hafta"),
        ("31.12.2026", "Aralık5.Hafta"),
    ],
)
def test_hafta_adi(tarih, beklenen):
    from core.tarih import hafta_adi

    assert hafta_adi(tarih) == beklenen


def test_hafta_adi_bozuk_tarihte_bugune_duser():
    from datetime import date

    from core.tarih import hafta_adi

    assert hafta_adi("olmayan tarih") == hafta_adi(date.today())
    assert hafta_adi("") == hafta_adi(date.today())


def test_hafta_adi_excel_sinirina_uyuyor():
    """Excel sayfa adı en fazla 31 karakter olabilir."""
    from core.tarih import AYLAR, hafta_adi

    for ay in range(1, 13):
        assert len(hafta_adi(f"15.{ay:02d}.2026")) <= 31
    assert all(a in "".join(AYLAR) for a in "Ağustos")


# --- Girdi doğrulaması -------------------------------------------------------


def test_bos_ad_soyad_reddedilir():
    veri = VardiyaVerisi(vardiya_adi="A", kayitlar=[VardiyaKaydi("", "Şef")])
    with pytest.raises(GirdiHatasi, match="Ad Soyad boş"):
        vardiya.uret(veri)


def test_bos_liste_uretilebilir():
    """Kayıt olmadan da geçerli bir dosya üretilmeli."""
    ws = _kaydet_ve_ac(vardiya.uret(VardiyaVerisi(vardiya_adi="B"))).active
    assert ws.cell(1, 1).value.startswith("VARDİYA LİSTESİ")
    assert ws.cell(vardiya.BASLIK_SATIRI, 2).value == "AD SOYAD"


# --- Üç vardiyalık yerleşim --------------------------------------------------


def test_uc_blok_yan_yana(haftalik):
    ws = _kaydet_ve_ac(vardiya.uret(haftalik)).active
    for blok, harf in enumerate("ABC"):
        bant = ws.cell(vardiya.VARDIYA_BANT_SATIRI, vardiya.blok_ilk_sutunu(blok))
        assert f"{harf} VARDİYASI" in bant.value
        basliki = ws.cell(vardiya.BASLIK_SATIRI, vardiya.blok_ilk_sutunu(blok) + 1)
        assert basliki.value == "AD SOYAD"


def test_her_blok_kendi_kayitlarini_tasiyor(haftalik):
    ws = _kaydet_ve_ac(vardiya.uret(haftalik)).active
    for blok, beklenen in enumerate(["A-Kisi", "B-Kisi", "C-Kisi"]):
        sutun = vardiya.blok_ilk_sutunu(blok) + 1
        assert ws.cell(vardiya.ILK_VERI_SATIRI, sutun).value == beklenen


def test_sayfa_adi_haftadan_turetiliyor(haftalik):
    assert _kaydet_ve_ac(vardiya.uret(haftalik)).active.title == "Ağustos3.Hafta"


def test_bloklar_ayni_sutun_genisliginde(haftalik):
    from openpyxl.utils import get_column_letter

    ws = _kaydet_ve_ac(vardiya.uret(haftalik)).active
    for i in range(len(vardiya.SUTUNLAR)):
        genislikler = {
            ws.column_dimensions[
                get_column_letter(vardiya.blok_ilk_sutunu(b) + i)
            ].width
            for b in range(3)
        }
        assert len(genislikler) == 1, f"{i}. sutun bloklar arasinda farkli."


def test_kisa_bloklar_bos_cerceveyle_tamamlaniyor(haftalik):
    """C vardiyasinda 1 kisi var; blok yine de A kadar uzun cercevelenmeli."""
    ws = _kaydet_ve_ac(vardiya.uret(haftalik)).active
    son = vardiya.ILK_VERI_SATIRI + haftalik.en_uzun_liste + vardiya.YEDEK_SATIR - 1
    h = ws.cell(son, vardiya.blok_ilk_sutunu(2) + 1)
    assert h.border.left.style == "thin", "Bos blok cercevesiz kalmis."


def test_tanimsiz_vardiya_harfi_reddedilir():
    """Sessiz veri kaybi olmamali: bilinmeyen harf hata vermeli."""
    from core.models import HaftalikVardiya

    with pytest.raises(GirdiHatasi, match="Tanınmayan vardiya adı"):
        HaftalikVardiya(vardiyalar=[VardiyaVerisi(vardiya_adi="D")])


def test_harfsiz_tek_vardiya_kaybolmuyor():
    """Harf verilmemis tek vardiya ilk bloga yerlesmeli, silinmemeli."""
    veri = VardiyaVerisi(kayitlar=[VardiyaKaydi("Tek Kisi", "Şef")])
    ws = _kaydet_ve_ac(vardiya.uret(veri)).active
    sutun = vardiya.blok_ilk_sutunu(0) + 1
    assert ws.cell(vardiya.ILK_VERI_SATIRI, sutun).value == "Tek Kisi"

"""FONKSİYON 1 — İş Talimatı üreticisinin otomatik doğrulama testleri.

Bu testlerin çekirdeği tek bir sorudur: şablonda var olan her çizim nesnesi
çıktıda da var mı? openpyxl bu nesneleri sessizce sildiği için bu denetim
projenin en kritik güvencesidir.
"""

from __future__ import annotations

import re

import pytest

from core.errors import GirdiHatasi
from core.generators import talimat
from core.models import KontrolAdimi, TalimatVerisi
from core.ooxml.drawing import DrawingPart
from core.ooxml.package import XlsxPackage
from core.units import (
    LOGO_HEIGHT_EMU,
    LOGO_WIDTH_EMU,
    STEP_IMAGE_HEIGHT_EMU,
    STEP_IMAGE_WIDTH_EMU,
)
from core import validate

DRAWING = "xl/drawings/drawing1.xml"
SHEET = "xl/worksheets/sheet1.xml"


def _cizim(pkg: XlsxPackage) -> DrawingPart:
    return DrawingPart(pkg.read_text(DRAWING))


# --- Çizim nesnesi korunumu (en kritik denetim) ------------------------------


def test_tum_metin_kutulari_ciktida_mevcut(ornek_talimat, sablon_talimat):
    sablon = _cizim(XlsxPackage.open(sablon_talimat))
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))

    assert len(cikti.find_all_shapes()) >= len(sablon.find_all_shapes()), (
        "Şablondaki metin kutularının bir kısmı çıktıda kaybolmuş."
    )


def test_dokuz_kontrol_adimi_basligi_mevcut(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    metinler = {a.text.strip() for a in cikti.find_all_shapes()}
    for n in range(1, 10):
        assert f"{n}. KONTROL ADIMI" in metinler, f"{n}. adım başlığı eksik."


def test_dokuz_cycle_kutusu_uretildi(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    metinler = {a.text.strip() for a in cikti.find_all_shapes()}
    for i in range(9):
        assert f"CYCLE: {3 + i} SN" in metinler, f"{i + 1}. adımın CYCLE kutusu eksik."


def test_klonlanan_cycle_kutusu_bicimi_kaynakla_ayni(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    kaynak = cikti.find_by_text("CYCLE: 3 SN")
    klon = cikti.find_by_text("CYCLE: 7 SN")

    assert kaynak is not None and klon is not None
    assert kaynak.fill_color == klon.fill_color
    for desen in (r"<a:ln>.*?</a:ln>", r"<a:bodyPr[^>]*/>", r'sz="\d+"'):
        a = re.search(desen, kaynak.xml, re.S)
        b = re.search(desen, klon.xml, re.S)
        assert (a and a.group(0)) == (b and b.group(0)), f"Biçim farkı: {desen}"


def test_klonlanan_kutularin_kimlikleri_benzersiz(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    idler = [a.shape_id for a in cikti.anchors]
    assert len(idler) == len(set(idler)), "Çizim nesnesi kimlikleri çakışıyor."


# --- Ölçüler (EMU cinsinden TAM eşitlik) -------------------------------------


def test_logo_olcusu_tam(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    logo = next(a for a in cikti.anchors if a.name == "Kurum Logosu")
    assert logo.ext == (LOGO_WIDTH_EMU, LOGO_HEIGHT_EMU)
    assert logo.ext == (1180800, 558000)          # 3,28 cm x 1,55 cm


@pytest.mark.parametrize("n", range(1, 10))
def test_adim_gorseli_olcusu_tam(ornek_talimat, sablon_talimat, n):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    g = next(a for a in cikti.anchors if a.name == f"Kontrol Adimi {n} Gorseli")
    assert g.ext == (STEP_IMAGE_WIDTH_EMU, STEP_IMAGE_HEIGHT_EMU)
    assert g.ext == (6120000, 1800000)            # 17 cm x 5 cm


def test_adim_gorselleri_onecellanchor(ornek_talimat, sablon_talimat):
    """Tam ölçü garantisi yalnızca oneCellAnchor ile mümkündür."""
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    for a in cikti.anchors:
        if "Kontrol Adimi" in a.name:
            assert a.kind == "oneCellAnchor"


# --- Z-sırası ----------------------------------------------------------------


def test_gorseller_metin_kutularinin_arkasinda(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    ilk_sp = next(i for i, a in enumerate(cikti.anchors) if a.shape_kind == "sp")
    for i, a in enumerate(cikti.anchors):
        if "Kontrol Adimi" in a.name:
            assert i < ilk_sp, (
                f"'{a.name}' metin kutularının üstünde kalıyor "
                f"(konum {i}, ilk kutu {ilk_sp})."
            )


# --- Sayfa yapısı korunumu ---------------------------------------------------


def test_birlesik_hucreler_degismedi(ornek_talimat, sablon_talimat):
    def merges(pkg):
        return set(re.findall(r'<mergeCell ref="([^"]+)"', pkg.read_text(SHEET)))

    assert merges(talimat.uret(ornek_talimat, sablon_talimat)) == merges(
        XlsxPackage.open(sablon_talimat)
    )


def test_baski_ayarlari_degismedi(ornek_talimat, sablon_talimat):
    a = XlsxPackage.open(sablon_talimat).read_text(SHEET)
    b = talimat.uret(ornek_talimat, sablon_talimat).read_text(SHEET)
    for etiket in ("pageSetup", "pageMargins", "printOptions"):
        assert re.search(rf"<{etiket}\b[^>]*/?>", a).group(0) == re.search(
            rf"<{etiket}\b[^>]*/?>", b
        ).group(0), f"<{etiket}> değişmiş."


# --- P sütunu sabitlemesi ----------------------------------------------------


def _col_widths(xml: str) -> dict[int, str]:
    """<cols> tanımlarını {sütun indeksi: genişlik} olarak açar."""
    genislikler = {}
    for m in re.finditer(r'<col min="(\d+)" max="(\d+)"[^>]*width="([\d.]+)"', xml):
        for c in range(int(m.group(1)), min(int(m.group(2)), 30) + 1):
            genislikler[c] = m.group(3)
    return genislikler


def test_p_sutunu_sabit_genislikte(ornek_talimat, sablon_talimat):
    b = talimat.uret(ornek_talimat, sablon_talimat).read_text(SHEET)
    assert _col_widths(b)[talimat.P_SUTUN_INDEKSI] == "26.57"


def test_diger_sutunlar_degismedi(ornek_talimat, sablon_talimat):
    """P dışındaki her sütun şablondaki genişliğini korumalı."""
    a = _col_widths(XlsxPackage.open(sablon_talimat).read_text(SHEET))
    b = _col_widths(talimat.uret(ornek_talimat, sablon_talimat).read_text(SHEET))

    assert set(a) == set(b), "Sütun tanımlarının kapsamı değişmiş."
    for sutun in a:
        if sutun == talimat.P_SUTUN_INDEKSI:
            continue
        assert a[sutun] == b[sutun], f"{sutun}. sütunun genişliği değişmiş."


def test_p_sutunu_ayrimi_komsulari_bozmuyor():
    """Ortak aralıkta tanımlı bir sütun bölünürken komşular korunmalı."""
    from core.ooxml.sheet import SheetPart

    sheet = SheetPart('<x><cols><col min="1" max="20" width="9.5" customWidth="1"/></cols></x>')
    sheet.sutun_genisligi_ayarla(16, 26.57)

    g = _col_widths(sheet.xml)
    assert g[16] == "26.57"
    assert g[15] == "9.5" and g[17] == "9.5"


def test_kritik_parcalar_korundu(ornek_talimat, sablon_talimat):
    sablon = XlsxPackage.open(sablon_talimat)
    cikti = talimat.uret(ornek_talimat, sablon_talimat)
    for ad in sablon.names():
        assert cikti.has(ad), f"Şablon parçası kayboldu: {ad}"


def test_sablon_dosyasi_degistirilmedi(ornek_talimat, sablon_talimat):
    """Şablon her zaman salt okunur kabul edilir."""
    once = sablon_talimat.read_bytes()
    talimat.uret(ornek_talimat, sablon_talimat)
    assert sablon_talimat.read_bytes() == once


# --- Hücre içerikleri --------------------------------------------------------


def test_konu_otomatik_eki(sablon_talimat):
    veri = TalimatVerisi(baslik="Test", konu="10598-AG")
    assert veri.konu_metni == "10598-AG İŞ TALİMATI HK."

    veri.konu_otomatik_ek = False
    assert veri.konu_metni == "10598-AG"


def test_turkce_karakterler_korunuyor(ornek_talimat, sablon_talimat):
    import openpyxl, io

    pkg = talimat.uret(ornek_talimat, sablon_talimat)
    tampon = io.BytesIO()
    import zipfile

    with zipfile.ZipFile(tampon, "w") as zf:
        for ad in pkg.names():
            zf.writestr(ad, pkg.read_bytes(ad))
    tampon.seek(0)

    ws = openpyxl.load_workbook(tampon).active
    assert ws["N6"].value == "Emirhan Koç"
    assert ws["R6"].value == "Şükrü Güngör"
    assert ws["C7"].value == "FIREWALL SAĞ"


def test_sari_alan_puntosu_ve_renkleri(ornek_talimat, sablon_talimat):
    pkg = talimat.uret(ornek_talimat, sablon_talimat)
    xml = pkg.read_text(SHEET)
    hucre = re.search(r'<c r="A20".*?</c>', xml, re.S).group(0)

    assert 'rgb="FFFF0000"' in hucre, "Başlık kırmızı değil."
    assert 'rgb="FF000000"' in hucre, "Açıklama siyah değil."
    assert hucre.count('<sz val="14"/>') >= 2, "Sarı alan 14 punto değil."
    assert "\n" in hucre, "Başlık ve açıklama alt alta değil."


def test_sari_alanda_siyah_metin_de_kalin(ornek_talimat, sablon_talimat):
    """Kullanıcı talebi: kırmızı başlığın yanı sıra siyah açıklama da kalın."""
    xml = talimat.uret(ornek_talimat, sablon_talimat).read_text(SHEET)
    hucre = re.search(r'<c r="A20".*?</c>', xml, re.S).group(0)

    siyah = re.search(
        r"<r><rPr>(?:(?!</r>).)*?FF000000(?:(?!</r>).)*?</rPr>", hucre, re.S
    )
    assert siyah is not None, "Siyah çalışma bulunamadı."
    assert "<b/>" in siyah.group(0), "Siyah açıklama kalın değil."

    # Her iki çalışma da kalın olmalı
    assert hucre.count("<b/>") >= 2


# --- İSG ikonları ------------------------------------------------------------


def test_isg_ikonlari_yerlestirildi(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    adlar = {a.name for a in cikti.find_pictures()}
    for ad in ["baret", "gozluk", "eldiven", "ayakkabi"]:
        assert f"ISG {ad}" in adlar, f"{ad} ikonu çıktıda yok."


def test_isg_ikonlari_esit_ve_kare(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    ikonlar = [a for a in cikti.anchors if a.name.startswith("ISG ")]

    olculer = {a.ext for a in ikonlar}
    assert len(olculer) == 1, "İkonlar farklı ölçülerde."
    cx, cy = olculer.pop()
    assert cx == cy, "İkonlar kare değil."


def test_isg_ikonlari_alana_sigiyor(ornek_talimat, sablon_talimat):
    from core.ooxml.layout import SheetLayout

    pkg = talimat.uret(ornek_talimat, sablon_talimat)
    layout = SheetLayout(pkg.read_text(SHEET))
    ikonlar = [a for a in _cizim(pkg).anchors if a.name.startswith("ISG ")]

    olcu = ikonlar[0].ext[0]
    toplam = len(ikonlar) * olcu + (len(ikonlar) - 1) * talimat.ISG_ARALIK
    assert toplam <= layout.cols_emu(talimat.ISG_ILK_SUTUN, talimat.ISG_SON_SUTUN)
    assert ikonlar[0].ext[1] <= layout.row_emu(talimat.ISG_SATIR)


def test_isg_ikonlari_satir_7de(ornek_talimat, sablon_talimat):
    cikti = _cizim(talimat.uret(ornek_talimat, sablon_talimat))
    for a in cikti.anchors:
        if a.name.startswith("ISG "):
            _, _, satir, _ = a.from_marker
            assert satir == talimat.ISG_SATIR, f"{a.name} yanlış satırda."


def test_ikon_secilmezse_sablona_dokunulmaz(sablon_talimat):
    """Hiç ikon seçilmezse şablonun piktogramları silinmemeli."""
    veri = TalimatVerisi(baslik="Test", isg_ikonlari=[])
    sablon = DrawingPart(XlsxPackage.open(sablon_talimat).read_text(DRAWING))
    cikti = _cizim(talimat.uret(veri, sablon_talimat))

    assert len(cikti.find_pictures()) == len(sablon.find_pictures())
    assert talimat.silinen_isg_gorseli(veri) == 0


def test_iki_ikon_secilirse_fazlasi_silinir(sablon_talimat):
    veri = TalimatVerisi(baslik="Test", isg_ikonlari=["baret", "eldiven"])
    assert talimat.silinen_isg_gorseli(veri) == 1

    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    assert len([a for a in cikti.anchors if a.name.startswith("ISG ")]) == 2


def test_bes_ikon_secilebilir(sablon_talimat):
    veri = TalimatVerisi(
        baslik="Test",
        isg_ikonlari=["baret", "gozluk", "kulaklik", "eldiven", "ayakkabi"],
    )
    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    assert len([a for a in cikti.anchors if a.name.startswith("ISG ")]) == 5


def test_sc_karakteristik_sembolu_degismiyor(ornek_talimat, sablon_talimat):
    """F6:G7'deki SC sembolü sabittir; kullanıcıdan istenmez, dokunulmaz."""
    sablon_pkg = XlsxPackage.open(sablon_talimat)
    cikti_pkg = talimat.uret(ornek_talimat, sablon_talimat)

    sc = next(
        a for a in DrawingPart(
            sablon_pkg.read_text(DRAWING)
        ).find_pictures() if a.name == "Resim 21"
    )
    cikti_sc = next(
        a for a in _cizim(cikti_pkg).find_pictures() if a.name == "Resim 21"
    )
    assert cikti_sc.xml == sc.xml, "SC sembolünün çizim tanımı değişmiş."

    hedef = "xl/media/image5.png"
    assert cikti_pkg.read_bytes(hedef) == sablon_pkg.read_bytes(hedef)


def test_isg_ikonu_secilince_metin_yazilmaz(sablon_talimat):
    """V7:W7 ikonlarla kaplanır; metin yazmak çakışma yaratırdı."""
    veri = TalimatVerisi(
        baslik="Test", isg_ekipmani="Eldiven, Gözlük", isg_ikonlari=["eldiven"]
    )
    xml = talimat.uret(veri, sablon_talimat).read_text(SHEET)
    assert "Eldiven, Gözlük" not in xml

    # İkon seçilmezse metin yazılır (geriye dönük uyum)
    veri2 = TalimatVerisi(baslik="Test", isg_ekipmani="Eldiven, Gözlük")
    assert "Eldiven, Gözlük" in talimat.uret(veri2, sablon_talimat).read_text(SHEET)


def test_gecersiz_ikon_yok_sayilir(sablon_talimat):
    veri = TalimatVerisi(baslik="Test", isg_ikonlari=["baret", "uzay_kaskı"])
    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    assert len([a for a in cikti.anchors if a.name.startswith("ISG ")]) == 1


# --- Kısmi doldurma ----------------------------------------------------------


def test_uc_adimla_uretim_cerceveli(sablon_talimat, gorsel_uret):
    veri = TalimatVerisi(baslik="Kısmi Test", bos_blok_davranisi="cerceveli")
    for i in range(3):
        veri.adimlar[i] = KontrolAdimi(
            baslik=f"{i + 1}. ADIM", cycle_sn=5, gorsel=gorsel_uret()
        )

    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    metinler = {a.text.strip() for a in cikti.find_all_shapes()}

    # Doldurulmayan blokların kutuları çerçeveli olarak YERİNDE kalmalı.
    for n in range(1, 10):
        assert f"{n}. KONTROL ADIMI" in metinler
    assert len([a for a in cikti.anchors if "Kontrol Adimi" in a.name]) == 3


def test_uc_adimla_uretim_temizle(sablon_talimat, gorsel_uret):
    veri = TalimatVerisi(baslik="Kısmi Test", bos_blok_davranisi="temizle")
    for i in range(3):
        veri.adimlar[i] = KontrolAdimi(
            baslik=f"{i + 1}. ADIM", cycle_sn=5, gorsel=gorsel_uret()
        )

    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    metinler = {a.text.strip() for a in cikti.find_all_shapes()}

    for n in range(1, 4):
        assert f"{n}. KONTROL ADIMI" in metinler
    for n in range(4, 10):
        assert f"{n}. KONTROL ADIMI" not in metinler, f"{n}. kutu temizlenmemiş."


def test_ilk_adim_bosken_cycle_kutusu_da_temizlenir(sablon_talimat, gorsel_uret):
    """1. adım boş bırakılırsa şablondan gelen CYCLE kutusu da kalmamalı."""
    veri = TalimatVerisi(baslik="Test", bos_blok_davranisi="temizle")
    veri.adimlar[4] = KontrolAdimi(baslik="5. ADIM", cycle_sn=7)

    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    metinler = {a.text.strip() for a in cikti.find_all_shapes()}

    assert talimat.SABLON_CYCLE_METNI not in metinler
    assert "CYCLE: 7 SN" in metinler
    assert "1. KONTROL ADIMI" not in metinler
    assert "5. KONTROL ADIMI" in metinler


def test_temizlenecek_kutular_uretimle_tutarli(sablon_talimat):
    """Doğrulayıcıya bildirilen sayı, gerçekte silinenle birebir aynı olmalı."""
    veri = TalimatVerisi(baslik="Test", bos_blok_davranisi="temizle")
    veri.adimlar[0] = KontrolAdimi(baslik="1. ADIM", cycle_sn=3)
    veri.adimlar[1] = KontrolAdimi(baslik="2. ADIM")

    sablon_sayi = len(
        DrawingPart(
            XlsxPackage.open(sablon_talimat).read_text(DRAWING)
        ).find_all_shapes()
    )
    cikti_sayi = len(_cizim(talimat.uret(veri, sablon_talimat)).find_all_shapes())

    assert sablon_sayi - cikti_sayi == len(talimat.temizlenecek_kutular(veri))


def test_cerceveli_modda_hicbir_kutu_silinmez(sablon_talimat):
    veri = TalimatVerisi(baslik="Test", bos_blok_davranisi="cerceveli")
    assert talimat.temizlenecek_kutular(veri) == []

    sablon = DrawingPart(XlsxPackage.open(sablon_talimat).read_text(DRAWING))
    cikti = _cizim(talimat.uret(veri, sablon_talimat))
    assert len(cikti.find_all_shapes()) == len(sablon.find_all_shapes())


def test_bos_sablon_birebir_kopya(sablon_talimat):
    kaynak = XlsxPackage.open(sablon_talimat)
    kopya = talimat.bos_sablon(sablon_talimat)
    assert kopya.names() == kaynak.names()
    for ad in kaynak.names():
        assert kopya.read_bytes(ad) == kaynak.read_bytes(ad)


# --- Girdi doğrulaması -------------------------------------------------------


def test_bos_baslik_reddedilir(sablon_talimat):
    with pytest.raises(GirdiHatasi, match="Talimat adı"):
        talimat.uret(TalimatVerisi(baslik="   "), sablon_talimat)


def test_asiri_uzun_sari_metin_reddedilir(sablon_talimat):
    veri = TalimatVerisi(baslik="Test")
    veri.adimlar[0] = KontrolAdimi(baslik="A" * 100, aciklama="B" * 200)
    with pytest.raises(GirdiHatasi, match="sığmıyor"):
        talimat.uret(veri, sablon_talimat)


def test_negatif_cycle_reddedilir(sablon_talimat):
    veri = TalimatVerisi(baslik="Test")
    veri.adimlar[0] = KontrolAdimi(baslik="A", cycle_sn=-1)
    with pytest.raises(GirdiHatasi, match="negatif"):
        talimat.uret(veri, sablon_talimat)


# --- Uçtan uca doğrulayıcı ---------------------------------------------------


def test_cikti_dogrulayicisi_temiz(ornek_talimat, sablon_talimat):
    pkg = talimat.uret(ornek_talimat, sablon_talimat)
    bulgular = validate.dogrula(
        pkg,
        sablon_talimat,
        beklenen_metinler=[f"{n}. KONTROL ADIMI" for n in range(1, 10)]
        + [f"CYCLE: {3 + i} SN" for i in range(9)],
        beklenen_gorsel_olculeri=[("Kurum Logosu", LOGO_WIDTH_EMU, LOGO_HEIGHT_EMU)]
        + [
            (f"Kontrol Adimi {i + 1} Gorseli", STEP_IMAGE_WIDTH_EMU, STEP_IMAGE_HEIGHT_EMU)
            for i in range(9)
        ],
    )
    assert bulgular == []


def test_openpyxl_referans_testi_basarisiz_olur(ornek_talimat, sablon_talimat, tmp_path):
    """Kontrol testi: openpyxl ile kaydetmek doğrulayıcıyı DÜŞÜRMELİ.

    Bu test doğrulayıcının gerçekten iş gördüğünü kanıtlar; geçerse
    doğrulayıcı sahte güven veriyor demektir.
    """
    import openpyxl

    from core.errors import DogrulamaHatasi

    yol = tmp_path / "openpyxl_ciktisi.xlsx"
    openpyxl.load_workbook(sablon_talimat).save(yol)

    with pytest.raises(DogrulamaHatasi):
        validate.dogrula(XlsxPackage.open(yol), sablon_talimat)


def test_asiri_buyuk_gorsel_reddedilir():
    """Düz renkli dev bir PNG diskte küçüktür ama bellekte yüzlerce MB ister.

    Bayt sınırı bunu yakalamaz; piksel sınırı yakalamalıdır. Yakalamazsa
    tek bir talimat isteği (logo + 9 adım fotoğrafı) fabrika makinesinin
    belleğini tüketebilir.
    """
    import io

    from PIL import Image

    from core import imaging
    from core.errors import GorselHatasi

    kenar = int((imaging.MAKS_PIKSEL * 1.5) ** 0.5)
    tampon = io.BytesIO()
    Image.new("RGB", (kenar, kenar), (10, 20, 30)).save(tampon, format="PNG")
    ham = tampon.getvalue()

    # Sınırı aşan görsel, bayt olarak küçük olsa bile reddedilmeli.
    assert len(ham) < imaging.MAKS_GORSEL_BOYUTU
    with pytest.raises(GorselHatasi) as hata:
        imaging.hazirla(ham, 1_000_000, 1_000_000)
    assert "piksel" in str(hata.value)


def test_makul_gorsel_kabul_edilir():
    """Sınır, gerçek bir fotoğraf makinesi çıktısını engellememelidir."""
    import io

    from PIL import Image

    from core import imaging

    tampon = io.BytesIO()
    Image.new("RGB", (4000, 3000), (10, 20, 30)).save(tampon, format="PNG")
    assert imaging.hazirla(tampon.getvalue(), 1_000_000, 1_000_000)

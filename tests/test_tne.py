"""FONKSİYON 2 — Tek Nokta Eğitimi üreticisinin otomatik doğrulama testleri."""

from __future__ import annotations

import re

import pytest

from core import validate
from core.errors import GirdiHatasi
from core.generators import tne
from core.models import TneKatilimci, TneVerisi
from core.ooxml.drawing import DrawingPart
from core.ooxml.package import XlsxPackage

DRAWING = "xl/drawings/drawing1.xml"
SHEET = "xl/worksheets/sheet1.xml"


@pytest.fixture
def ornek_tne(gorsel_uret):
    return TneVerisi(
        baslik="KAYNAK DİKİŞİ GÖZLE KONTROL EĞİTİMİ",
        konu="10598-AG",
        mudurluk_birim="Üretim Müdürlüğü / Pres Hattı",
        kisim="Kalite Kontrol",
        hazirlayan="Emirhan Koç",
        parametre="Kaynak dikiş boyu",
        olcum_araci="Kumpas",
        parca_etkisi="Mukavemet kaybı",
        sorumlu="Vardiya Şefi",
        egitim_suresi="15 DK",
        egitim_tarihi="17.08.2026",
        tnd_no="TND-2026-047",
        egitim_veren="Şükrü Güngör",
        talimat_no="BC-F-1328",
        tarih="17.08.2026",
        egitim_icerigi=["KALİTE", "ÜRETİM", "STANDART"],
        egitim_turu=["TEMEL BİLGİ", "HATA"],
        logo=gorsel_uret(600, 280),
        egitim_gorseli=gorsel_uret(1600, 900),
        katilimcilar=[
            TneKatilimci("Ahmet Yılmaz", "10234"),
            TneKatilimci("Ayşe Çelik", "10235"),
        ],
    )


def _cizim(pkg: XlsxPackage) -> DrawingPart:
    return DrawingPart(pkg.read_text(DRAWING))


# --- Çizim nesnesi korunumu --------------------------------------------------


def test_tum_cizim_nesneleri_korundu(ornek_tne, sablon_tne):
    sablon = _cizim(XlsxPackage.open(sablon_tne))
    cikti = _cizim(tne.uret(ornek_tne, sablon_tne))
    assert len(cikti.find_all_shapes()) == len(sablon.find_all_shapes())


def test_etiket_metinleri_bozulmadi(ornek_tne, sablon_tne):
    cikti = _cizim(tne.uret(ornek_tne, sablon_tne))
    metinler = {a.text.strip() for a in cikti.find_all_shapes()}
    for etiket in tne.EGITIM_ICERIGI + tne.EGITIM_TURU:
        assert etiket in metinler, f"'{etiket}' etiketi kaybolmuş."


# --- Onay kutuları (çizim nesnesi dolgusu) -----------------------------------


def test_secili_kutular_yesil(ornek_tne, sablon_tne):
    kutular = tne.onay_kutulari(_cizim(tne.uret(ornek_tne, sablon_tne)))
    for ad in ["KALİTE", "ÜRETİM", "STANDART", "TEMEL BİLGİ", "HATA"]:
        assert kutular[ad].fill_color == tne.ISARETLI_RENK, f"{ad} işaretli değil."


def test_secilmeyen_kutular_bos(ornek_tne, sablon_tne):
    kutular = tne.onay_kutulari(_cizim(tne.uret(ornek_tne, sablon_tne)))
    for ad in ["GÜVENLİK", "BAKIM", "ÇEVRE", "İYİLEŞTİRME"]:
        assert kutular[ad].fill_color is None, f"{ad} yanlışlıkla işaretlenmiş."


def test_sablondaki_isaretler_temizleniyor(sablon_tne, gorsel_uret):
    """Şablonda KALİTE/TEMEL BİLGİ/HATA işaretli gelir; seçilmezse temizlenmeli."""
    veri = TneVerisi(baslik="Test", egitim_icerigi=["BAKIM"], egitim_turu=[])
    kutular = tne.onay_kutulari(_cizim(tne.uret(veri, sablon_tne)))

    assert kutular["BAKIM"].fill_color == tne.ISARETLI_RENK
    for ad in ["KALİTE", "TEMEL BİLGİ", "HATA"]:
        assert kutular[ad].fill_color is None, f"{ad} şablondan kalmış."


def test_dokuz_onay_kutusu_tanindi(sablon_tne):
    kutular = tne.onay_kutulari(
        _cizim(XlsxPackage.open(sablon_tne))
    )
    assert set(kutular) == set(tne.EGITIM_ICERIGI) | set(tne.EGITIM_TURU)
    assert len(kutular) == 9


def test_kutular_benzersiz_eslesti(sablon_tne):
    """Her etiket FARKLI bir dikdörtgene eşleşmeli."""
    kutular = tne.onay_kutulari(_cizim(XlsxPackage.open(sablon_tne)))
    idler = [k.shape_id for k in kutular.values()]
    assert len(idler) == len(set(idler)), "İki etiket aynı kutuya eşleşmiş."


# --- Görseller ---------------------------------------------------------------


def test_egitim_gorseli_arkada(ornek_tne, sablon_tne):
    cikti = _cizim(tne.uret(ornek_tne, sablon_tne))
    konum = next(
        i for i, a in enumerate(cikti.anchors) if a.name == "Egitim Gorseli"
    )
    ilk_sp = next(i for i, a in enumerate(cikti.anchors) if a.shape_kind == "sp")
    assert konum < ilk_sp, "Eğitim görseli çizim nesnelerinin üstünde."


def test_egitim_gorseli_olcusu_tam(ornek_tne, sablon_tne):
    """Ölçü SABİTTİR: 41,75 x 23,72 cm — görselin oranından bağımsız."""
    from core.units import TNE_IMAGE_HEIGHT_EMU, TNE_IMAGE_WIDTH_EMU

    cikti = _cizim(tne.uret(ornek_tne, sablon_tne))
    g = next(a for a in cikti.anchors if a.name == "Egitim Gorseli")

    assert g.ext == (TNE_IMAGE_WIDTH_EMU, TNE_IMAGE_HEIGHT_EMU)
    assert g.ext == (15030000, 8539200)
    assert g.kind == "oneCellAnchor", "Tam ölçü yalnızca oneCellAnchor'da garanti."


@pytest.mark.parametrize("boyut", [(1600, 900), (900, 1600), (1000, 1000)])
def test_olcu_kaynak_oranindan_bagimsiz(sablon_tne, gorsel_uret, boyut):
    """Dikey, yatay veya kare — çıktı kutusu hep aynı ölçüde olmalı."""
    from core.units import TNE_IMAGE_HEIGHT_EMU, TNE_IMAGE_WIDTH_EMU

    veri = TneVerisi(egitim_gorseli=gorsel_uret(*boyut))
    g = next(
        a for a in _cizim(tne.uret(veri, sablon_tne)).anchors
        if a.name == "Egitim Gorseli"
    )
    assert g.ext == (TNE_IMAGE_WIDTH_EMU, TNE_IMAGE_HEIGHT_EMU)


def test_gorsel_kutuyu_bosluksuz_doldurur(sablon_tne, gorsel_uret):
    """Görsel kutuyu TAMAMEN doldurur; oran korunur, taşan kısım kırpılır.

    Önceki davranış sığdırmaktı ve artan alanı beyaz bırakıyordu: kare bir
    kaynak, geniş TNE kutusunda yanlarda geniş beyaz bantlar üretiyordu.
    Ürün kararı doldurma yönünde değişti (bkz. core/imaging.py).
    """
    import io

    from PIL import Image

    from core.units import TNE_IMAGE_HEIGHT_EMU, TNE_IMAGE_WIDTH_EMU

    # Kare kaynak, geniş kutu: eskiden yanlar beyaz kalıyordu.
    veri = TneVerisi(egitim_gorseli=gorsel_uret(600, 600, renk=(0, 0, 0)))
    pkg = tne.uret(veri, sablon_tne)

    medya = next(a for a in pkg.names() if a.startswith("xl/media/image2"))
    with Image.open(io.BytesIO(pkg.read_bytes(medya))) as img:
        hedef_oran = TNE_IMAGE_WIDTH_EMU / TNE_IMAGE_HEIGHT_EMU
        assert abs(img.width / img.height - hedef_oran) < 0.01

        # Kenarlar da kaynak görselden gelmeli: hiçbir yerde beyaz dolgu yok.
        for x in (2, img.width // 2, img.width - 3):
            assert img.getpixel((x, img.height // 2)) != (255, 255, 255), (
                f"x={x} noktasında beyaz dolgu var — kutu doldurulmamış"
            )


def test_logo_olcusu_sablondan_devralindi(ornek_tne, sablon_tne):
    """Faz 0 kararı: TNE logosu şablondaki ölçüsünü korur."""
    sablon_logo = next(
        a for a in _cizim(XlsxPackage.open(sablon_tne)).find_pictures()
    )
    cikti_logo = next(
        a for a in _cizim(tne.uret(ornek_tne, sablon_tne)).anchors
        if a.name == "Kurum Logosu"
    )
    assert cikti_logo.ext == sablon_logo.ext


def test_medya_bicimi_uzantiyla_uyumlu(ornek_tne, sablon_tne):
    """Şablonun logosu .jpeg'dir; oraya PNG baytı yazılmamalı."""
    pkg = tne.uret(ornek_tne, sablon_tne)
    veri = pkg.read_bytes("xl/media/image1.jpeg")
    assert veri[:2] == b"\xff\xd8", "JPEG parçasına JPEG olmayan veri yazılmış."


# --- Hücre içerikleri --------------------------------------------------------


def test_baslik_24_punto(ornek_tne, sablon_tne):
    """2.2 — Şablonda 36 punto, talep gereği 24 punto uygulanır."""
    import io
    import zipfile

    import openpyxl

    pkg = tne.uret(ornek_tne, sablon_tne)
    tampon = io.BytesIO()
    with zipfile.ZipFile(tampon, "w") as zf:
        for ad in pkg.names():
            zf.writestr(ad, pkg.read_bytes(ad))
    tampon.seek(0)

    ws = openpyxl.load_workbook(tampon).worksheets[0]
    assert ws["B2"].font.sz == 24
    assert ws["C7"].font.sz == 14
    assert ws["G7"].font.sz == 10


def test_konu_otomatik_eki():
    veri = TneVerisi(baslik="T", konu="10598-AG")
    assert veri.konu_metni == "10598-AG TEK NOKTA EĞİTİMİ HK."
    veri.konu_otomatik_ek = False
    assert veri.konu_metni == "10598-AG"


def test_etiket_onekleri_korundu(ornek_tne, sablon_tne):
    """Etiket ve değer aynı hücrede olan alanlarda etiket bozulmamalı."""
    xml = tne.uret(ornek_tne, sablon_tne).read_text(SHEET)
    for parca in ("MÜDÜRLÜK / BİRİM :", "KISIM :", "HAZIRLAYAN :", "SAYFA NO"):
        assert parca in xml, f"'{parca}' etiketi kaybolmuş."


def test_katilimci_imza_sutunu_bos(ornek_tne, sablon_tne):
    """2.7 — İmza sütunu ıslak imza için boş bırakılır."""
    xml = tne.uret(ornek_tne, sablon_tne).read_text(SHEET)
    for satir in (11, 12):
        hucre = re.search(rf'<c r="J{satir}"[^>]*/>', xml)
        assert hucre is not None, f"J{satir} doldurulmuş olmamalı."


# --- Sayfa yapısı ------------------------------------------------------------


def test_birlesik_hucreler_ve_baski_ayarlari(ornek_tne, sablon_tne):
    a = XlsxPackage.open(sablon_tne).read_text(SHEET)
    b = tne.uret(ornek_tne, sablon_tne).read_text(SHEET)

    assert set(re.findall(r'<mergeCell ref="([^"]+)"', a)) == set(
        re.findall(r'<mergeCell ref="([^"]+)"', b)
    )
    for etiket in ("pageSetup", "pageMargins", "printOptions", "cols"):
        assert re.search(rf"<{etiket}\b[^>]*/?>", a).group(0) == re.search(
            rf"<{etiket}\b[^>]*/?>", b
        ).group(0)


def test_yardimci_sayfa_korundu(ornek_tne, sablon_tne):
    """Boş 'Sayfa1' yardımcı sayfası silinmemeli."""
    assert tne.uret(ornek_tne, sablon_tne).has("xl/worksheets/sheet2.xml")


def test_sablon_dosyasi_degistirilmedi(ornek_tne, sablon_tne):
    once = sablon_tne.read_bytes()
    tne.uret(ornek_tne, sablon_tne)
    assert sablon_tne.read_bytes() == once


def test_bos_sablon_birebir_kopya(sablon_tne):
    kaynak = XlsxPackage.open(sablon_tne)
    kopya = tne.bos_sablon(sablon_tne)
    for ad in kaynak.names():
        assert kopya.read_bytes(ad) == kaynak.read_bytes(ad)


def test_cikti_dogrulayicisi_temiz(ornek_tne, sablon_tne):
    assert validate.dogrula(
        tne.uret(ornek_tne, sablon_tne),
        sablon_tne,
        beklenen_metinler=tne.EGITIM_ICERIGI + tne.EGITIM_TURU,
    ) == []


# --- Girdi doğrulaması -------------------------------------------------------


def test_baslik_zorunlu_degil(sablon_tne):
    """Kullanıcıdan başlık istenmez; şablon kendi başlığıyla gelir."""
    pkg = tne.uret(TneVerisi(egitim_suresi="15 DK"), sablon_tne)
    assert "TEK NOKTA EĞİTİMİ" in pkg.read_text("xl/sharedStrings.xml")


def test_sadece_istenen_alanlar_yeterli(sablon_tne, gorsel_uret):
    """Arayüzün gönderdiği asgari veri seti tek başına geçerli olmalı."""
    veri = TneVerisi(
        egitim_icerigi=["KALİTE"],
        egitim_turu=["TEMEL BİLGİ"],
        egitim_suresi="15 DK",
        sorumlu="Vardiya Şefi",
        egitim_veren="Şükrü Güngör",
        egitim_tarihi="17.08.2026",
        egitim_gorseli=gorsel_uret(1600, 900),
    )
    pkg = tne.uret(veri, sablon_tne)
    assert validate.dogrula(pkg, sablon_tne) == []

    kutular = tne.onay_kutulari(_cizim(pkg))
    assert kutular["KALİTE"].fill_color == tne.ISARETLI_RENK
    assert any(a.name == "Egitim Gorseli" for a in _cizim(pkg).anchors)


def test_sorulmayan_alanlar_bos_kalir(sablon_tne):
    """Sahada kalemle doldurulacak alanlara program değer yazmamalı."""
    veri = TneVerisi(egitim_suresi="15 DK", sorumlu="Şef")
    xml = tne.uret(veri, sablon_tne).read_text("xl/worksheets/sheet1.xml")

    # Katılımcı satırları (11-42) tamamen boş kalmalı
    for satir in (11, 20, 42):
        for sutun in ("H", "I", "J"):
            hucre = re.search(rf'<c r="{sutun}{satir}"[^>]*/>', xml)
            assert hucre is not None, f"{sutun}{satir} doldurulmuş."


def test_gecersiz_egitim_icerigi_reddedilir(sablon_tne):
    with pytest.raises(GirdiHatasi, match="geçerli bir eğitim içeriği değil"):
        tne.uret(TneVerisi(baslik="T", egitim_icerigi=["MONTAJ"]), sablon_tne)


def test_asiri_katilimci_reddedilir(sablon_tne):
    veri = TneVerisi(
        baslik="T",
        katilimcilar=[TneKatilimci(f"Kişi {i}", str(i)) for i in range(40)],
    )
    with pytest.raises(GirdiHatasi, match="en fazla 32"):
        tne.uret(veri, sablon_tne)


# --- Dolgu ile KENARLIK ayrimi (regresyon) -----------------------------------
# Sablondaki bazi onay kutularinin dolgusu tema rengiyle (schemeClr bg1)
# tanimlidir ve kenarliklarinin kendi <a:solidFill> blogu vardir. Tum spPr
# uzerinde arama yapan bir uygulama, dolgu yerine CERCEVEYI boyar.


def _spPr(anchor) -> str:
    return re.search(r"<xdr:spPr>.*?</xdr:spPr>", anchor.xml, re.S).group(0)


def _kenarlik(anchor) -> str:
    sp = _spPr(anchor)
    return "<a:ln" + sp.split("<a:ln")[1] if "<a:ln" in sp else ""


def _dolgu_bloklari(anchor) -> int:
    """spPr'nin kenarlik ONCESI bolgesindeki dolgu tanimi sayisi."""
    bas = _spPr(anchor).split("<a:ln")[0]
    return bas.count("<a:solidFill>") + bas.count("<a:noFill/>")


def test_fill_color_kenarlik_rengini_okumaz(sablon_tne):
    """Kenarligi siyah, dolgusu tema beyazi olan kutu dolu sayilmamali."""
    d = _cizim(XlsxPackage.open(sablon_tne))
    kutular = tne.onay_kutulari(d)

    iyilestirme = kutular["İYİLEŞTİRME"]
    assert "<a:srgbClr val=\"000000\"/>" in _kenarlik(iyilestirme), (
        "Test dayanagi kayboldu: bu kutunun kenarligi siyah olmali."
    )
    assert iyilestirme.fill_color is None, (
        "Kenarlik rengi dolgu olarak okunuyor."
    )


def test_set_fill_kenarliga_dokunmaz(ornek_tne, sablon_tne):
    """Isaretleme yalnizca dolguyu degistirmeli; cerceve aynen kalmali."""
    sablon = {a.shape_id: a for a in _cizim(XlsxPackage.open(sablon_tne)).anchors}
    veri = TneVerisi(egitim_icerigi=["KALİTE"], egitim_turu=["İYİLEŞTİRME", "HATA"])
    cikti = _cizim(tne.uret(veri, sablon_tne))

    for a in cikti.anchors:
        if a.shape_kind != "sp" or a.text.strip():
            continue
        assert _kenarlik(a) == _kenarlik(sablon[a.shape_id]), (
            f"id={a.shape_id} kutusunun kenarligi degismis."
        )


def test_her_kutuda_tek_dolgu_tanimi(sablon_tne):
    """Eski dolgu silinmeden yenisi eklenirse gecersiz XML olusur."""
    veri = TneVerisi(egitim_icerigi=["KALİTE"], egitim_turu=["İYİLEŞTİRME", "HATA"])
    for a in _cizim(tne.uret(veri, sablon_tne)).anchors:
        if a.shape_kind != "sp" or a.text.strip():
            continue
        assert _dolgu_bloklari(a) == 1, (
            f"id={a.shape_id} kutusunda {_dolgu_bloklari(a)} dolgu tanimi var."
        )


def test_tema_dolgulu_kutu_isaretlenebiliyor(sablon_tne):
    """İYİLEŞTİRME kutusunun dolgusu schemeClr'dir; yine de yesillenmeli."""
    veri = TneVerisi(egitim_turu=["İYİLEŞTİRME"])
    kutular = tne.onay_kutulari(_cizim(tne.uret(veri, sablon_tne)))

    assert kutular["İYİLEŞTİRME"].fill_color == tne.ISARETLI_RENK
    assert "schemeClr" not in _spPr(kutular["İYİLEŞTİRME"]).split("<a:ln")[0], (
        "Eski tema dolgusu silinmemis."
    )

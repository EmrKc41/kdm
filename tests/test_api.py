"""Arayüz API'sinin uçtan uca testleri.

Tarayıcının gönderdiği gövdenin aynısı gönderilir; dönen baytın gerçekten
açılabilir bir xlsx olduğu ve çizim nesnelerini taşıdığı doğrulanır.
"""

from __future__ import annotations

import base64
import io
import zipfile

import pytest

pytest.importorskip("fastapi")

from core.ooxml.drawing import DrawingPart      # noqa: E402
from tests.asgi_istemci import AsgiIstemci      # noqa: E402
from tests.conftest import sablon_gerekir      # noqa: E402
from ui.app import app                          # noqa: E402

istemci = AsgiIstemci(app)


@pytest.fixture(autouse=True)
def _oturum():
    """Aşağıdaki testler giriş yapmış bir arayüzü taklit eder.

    Motor artık /api/* için oturum çerezi ister; oturumsuz davranış ayrıca
    test_oturum.py'de sınanır. Burada her testin başında çerezin durduğundan
    emin oluruz, böylece testlerin sırası birbirini etkilemez.
    """
    if not istemci.cerezler:
        assert istemci.giris().status_code == 200
    yield

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _data_url(veri: bytes) -> str:
    return "data:image/png;base64," + base64.b64encode(veri).decode("ascii")


def _cizim(icerik: bytes) -> DrawingPart:
    with zipfile.ZipFile(io.BytesIO(icerik)) as zf:
        return DrawingPart(zf.read("xl/drawings/drawing1.xml").decode("utf-8"))


def _xlsx_mi(yanit) -> None:
    assert yanit.status_code == 200, yanit.text[:400]
    assert yanit.headers["content-type"] == XLSX_MIME
    assert zipfile.is_zipfile(io.BytesIO(yanit.content))


# --- Sayfa -------------------------------------------------------------------


def test_anasayfa_yukleniyor():
    yanit = istemci.get("/")
    assert yanit.status_code == 200
    assert "Kalite Doküman Merkezi" in yanit.text
    assert "localhost:3000" in yanit.text


def test_anasayfa_uzak_kaynak_icermez():
    """İnternetsiz çalışma: CDN / uzak font referansı olmamalı (localhost hariç)."""
    sayfa = istemci.get("/").text.lower()
    for yasak in ("cdnjs", "googleapis", "unpkg", "jsdelivr", "cloudflare.com"):
        assert yasak not in sayfa, f"Sayfada uzak kaynak referansı var: {yasak}"


# --- FONKSİYON 1 -------------------------------------------------------------


@sablon_gerekir
def test_talimat_uretimi(gorsel_uret):
    govde = {
        "baslik": "FIREWALL İŞ TALİMATI",
        "konu": "10598-AG",
        "konu_otomatik_ek": True,
        "parca_no": "10598-AG",
        "musteri": "ABC OTOMOTİV",
        "hazirlayan": "Emirhan Koç",
        "tarih": "17.08.2026",
        "logo": _data_url(gorsel_uret(600, 280)),
        "bos_blok_davranisi": "cerceveli",
        "adimlar": [
            {
                "baslik": f"{i + 1}. ADIM",
                "aciklama": "Gözle kontrol edilir.",
                "cycle_sn": str(3 + i),
                "gorsel": _data_url(gorsel_uret(1400, 400)),
            }
            for i in range(9)
        ],
    }
    yanit = istemci.post("/api/talimat", json=govde)
    _xlsx_mi(yanit)
    assert "10598_AG_IS_TALIMATI" in yanit.headers["content-disposition"]

    cizim = _cizim(yanit.content)
    metinler = {a.text.strip() for a in cizim.find_all_shapes()}
    for n in range(1, 10):
        assert f"{n}. KONTROL ADIMI" in metinler
        assert f"CYCLE: {2 + n} SN" in metinler


@sablon_gerekir
def test_talimat_kismi_doldurma(gorsel_uret):
    govde = {
        "baslik": "Kısmi",
        "adimlar": [{"baslik": "1. ADIM", "cycle_sn": "5"}] + [{} for _ in range(8)],
        "bos_blok_davranisi": "temizle",
    }
    yanit = istemci.post("/api/talimat", json=govde)
    _xlsx_mi(yanit)

    metinler = {a.text.strip() for a in _cizim(yanit.content).find_all_shapes()}
    assert "1. KONTROL ADIMI" in metinler
    assert "5. KONTROL ADIMI" not in metinler


def test_talimat_bos_baslik_400():
    yanit = istemci.post("/api/talimat", json={"baslik": ""})
    assert yanit.status_code == 400
    assert "Talimat adı" in yanit.json()["hata"]


def test_hata_yanitinda_stack_trace_yok():
    """Kullanıcıya asla teknik hata izi gösterilmez."""
    yanit = istemci.post("/api/talimat", json={"baslik": ""})
    govde = yanit.text
    for iz in ("Traceback", "File \"", "line ", ".py\""):
        assert iz not in govde, f"Yanıtta teknik iz sızmış: {iz}"


# --- FONKSİYON 2 -------------------------------------------------------------


@sablon_gerekir
def test_tne_uretimi(gorsel_uret):
    govde = {
        "baslik": "KAYNAK KONTROL EĞİTİMİ",
        "konu": "10598-AG",
        "mudurluk_birim": "Üretim Müdürlüğü",
        "egitim_icerigi": ["KALİTE", "ÜRETİM"],
        "egitim_turu": ["TEMEL BİLGİ"],
        "logo": _data_url(gorsel_uret(600, 280)),
        "egitim_gorseli": _data_url(gorsel_uret(1600, 900)),
        "katilimcilar": [{"ad_soyad": "Ayşe Çelik", "sicil_no": "10235"}],
        "tarih": "17.08.2026",
    }
    yanit = istemci.post("/api/tne", json=govde)
    _xlsx_mi(yanit)
    assert "TEK_NOKTA_EGITIMI" in yanit.headers["content-disposition"]

    from core.generators import tne

    kutular = tne.onay_kutulari(_cizim(yanit.content))
    assert kutular["KALİTE"].fill_color == "00B050"
    assert kutular["ÜRETİM"].fill_color == "00B050"
    assert kutular["HATA"].fill_color is None       # şablondan kalmamalı


def test_tne_gecersiz_secim_400():
    yanit = istemci.post(
        "/api/tne", json={"baslik": "T", "egitim_icerigi": ["MONTAJ"]}
    )
    assert yanit.status_code == 400
    assert "eğitim içeriği" in yanit.json()["hata"]


# --- FONKSİYON 3 -------------------------------------------------------------


def test_vardiya_uretimi():
    govde = {
        "vardiya_adi": "A VARDİYASI",
        "tarih": "17.08.2026",
        "vardiya_saati": "08:00 - 16:00",
        "normal_unvanlar": ["Kalite Operatörü"],
        "kayitlar": [
            {"ad_soyad": "Ahmet Yılmaz", "unvan": "Kalite Operatörü",
             "telefon": "0532 111 22 33", "durak": "Merkez"},
            {"ad_soyad": "Ayşe Çelik", "unvan": "Vardiya Formeni",
             "telefon": "0533 222 33 44", "durak": "Şehitler"},
        ],
    }
    yanit = istemci.post("/api/vardiya", json=govde)
    _xlsx_mi(yanit)

    import openpyxl

    from core.generators import vardiya as vg

    ws = openpyxl.load_workbook(io.BytesIO(yanit.content)).active
    assert ws.cell(vg.ILK_VERI_SATIRI, 2).font.bold is False
    assert ws.cell(vg.ILK_VERI_SATIRI + 1, 2).font.bold is True
    assert ws.cell(vg.ILK_VERI_SATIRI, 5).value == "0532 111 22 33"


def test_vardiya_csv_ice_aktarma():
    csv = (
        "Ad Soyad;Ünvan;Çalışacağı Yer;Telefon No;Durak İsmi\n"
        "Ahmet Yılmaz;Kalite Operatörü;Pres 3;0532 111 22 33;Merkez\n"
    ).encode("utf-8")

    yanit = istemci.post(
        "/api/vardiya/ice-aktar",
        json={
            "dosya_adi": "liste.csv",
            "icerik": "data:text/csv;base64," + base64.b64encode(csv).decode(),
        },
    )
    assert yanit.status_code == 200
    kayitlar = yanit.json()["kayitlar"]
    assert len(kayitlar) == 1
    assert kayitlar[0]["ad_soyad"] == "Ahmet Yılmaz"
    assert kayitlar[0]["telefon"] == "0532 111 22 33"


def test_ice_aktarma_bozuk_dosya_400():
    yanit = istemci.post(
        "/api/vardiya/ice-aktar",
        json={"dosya_adi": "liste.csv",
              "icerik": "data:text/csv;base64," + base64.b64encode(b"a;b\n").decode()},
    )
    assert yanit.status_code == 400
    assert "Ad Soyad" in yanit.json()["hata"]


# --- Boş şablonlar -----------------------------------------------------------


@pytest.mark.parametrize(
    "tip",
    [
        # talimat ve tne şablondan türetilir; vardiya ve rapor sıfırdan üretilir.
        pytest.param("talimat", marks=sablon_gerekir),
        pytest.param("tne", marks=sablon_gerekir),
        "vardiya",
        "rapor",
    ],
)
def test_bos_sablon_indirme(tip):
    _xlsx_mi(istemci.get(f"/api/bos/{tip}"))


@sablon_gerekir
def test_bos_talimat_cizim_nesnelerini_tasiyor():
    cizim = _cizim(istemci.get("/api/bos/talimat").content)
    assert len(cizim.find_all_shapes()) == 10


def test_bilinmeyen_tip_400():
    yanit = istemci.get("/api/bos/belge")
    assert yanit.status_code == 400
    assert "tanınmayan" in yanit.json()["hata"]


def test_rapor_uretimi():
    govde = {
        "baslik": "KALİTE UYGUNSUZLUK TAKİP RAPORU",
        "konu": "10598-AG",
        "rapor_no": "R-001",
        "tarih": "17.08.2026",
        "hazirlayan": "Test Kullanıcı",
        "genel_durum": "Açık",
        "ozet": "Test özeti",
        "satirlar": [
            {
                "tanim": "Test uygunsuzluk",
                "kok_neden": "Test neden",
                "duzeltici_faaliyet": "Test faaliyet",
                "sorumlu": "Kalite",
                "hedef_tarih": "20.08.2026",
                "durum": "Açık",
            }
        ],
    }
    _xlsx_mi(istemci.post("/api/rapor", json=govde))


@sablon_gerekir
def test_saglik_kontrolu():
    yanit = istemci.get("/api/health")
    assert yanit.status_code == 200
    govde = yanit.json()
    assert govde["durum"] == "acik"
    assert govde["sablonlar"]["talimat"] is True
    assert govde["sablonlar"]["tne"] is True
    assert govde["tum_sablonlar_hazir"] is True


# --- Ayarlar -----------------------------------------------------------------


def test_unvan_ayarlari_kaydet_oku():
    yeni = ["Kalite Operatörü", "Stajyer"]
    yanit = istemci.post("/api/ayarlar/unvanlar", json={"normal_unvanlar": yeni})
    assert yanit.status_code == 200
    assert yanit.json()["normal_unvanlar"] == yeni

    assert istemci.get("/api/ayarlar/unvanlar").json()["normal_unvanlar"] == yeni

    # Testin yan etkisini geri al
    istemci.post("/api/ayarlar/unvanlar", json={"normal_unvanlar": ["Kalite Operatörü"]})


def test_bos_unvan_listesi_reddedilir():
    yanit = istemci.post("/api/ayarlar/unvanlar", json={"normal_unvanlar": []})
    assert yanit.status_code == 400
    assert "en az bir ünvan" in yanit.json()["hata"].lower()


def test_kurallar_oku_kaydet():
    yanit = istemci.get("/api/ayarlar/kurallar")
    assert yanit.status_code == 200
    assert "kurallar" in yanit.json()

    govde = {
        "varsayilan": {"kalin": False, "renk": "000000", "italik": False},
        "kurallar": [
            {
                "ad": "Test kural",
                "alan": "unvan",
                "operator": "listede_degil",
                "degerler": ["Kalite Operatörü"],
                "bicim": {"kalin": True, "renk": "C00000", "italik": False},
                "etkin": True,
            }
        ],
    }
    assert istemci.post("/api/ayarlar/kurallar", json=govde).status_code == 200
    # Varsayılana geri al
    istemci.post("/api/ayarlar/unvanlar", json={"normal_unvanlar": ["Kalite Operatörü"]})
